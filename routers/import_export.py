import csv
import io
import os
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Request, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse
import database as db
from config import templates

router = APIRouter(prefix="/import-export", tags=["import-export"])


# ─── Plan normalisation ───────────────────────────────────────────────────────
# Accepts: "1 Month", "1 Months", "Monthly", "3 Months", "Quarterly", etc.
_PLAN_MONTHS = {"Monthly": 1, "Quarterly": 3, "Half Yearly": 6, "Annual": 12}

def _normalise_plan(raw: str) -> str:
    r = raw.strip().lower()
    # Numeric-month patterns: "1 month", "3 months", "6 months", "12 months"
    import re
    m = re.match(r"^(\d+)\s*months?$", r)
    if m:
        n = int(m.group(1))
        if n <= 1:  return "Monthly"
        if n <= 3:  return "Quarterly"
        if n <= 6:  return "Half Yearly"
        return "Annual"
    # Named variants
    if r in ("monthly",):                           return "Monthly"
    if r in ("quarterly", "3 monthly"):             return "Quarterly"
    if r in ("half yearly", "half-yearly", "6 monthly"): return "Half Yearly"
    if r in ("annual", "yearly", "12 monthly"):     return "Annual"
    return "Monthly"  # default


# ─── Date parsing ─────────────────────────────────────────────────────────────
_DATE_FMTS = [
    "%d-%b-%y",   # 1-Jan-26  ← user's format
    "%d-%b-%Y",   # 1-Jan-2026
    "%d/%b/%y",   # 1/Jan/26
    "%d/%b/%Y",   # 1/Jan/2026
    "%Y-%m-%d",   # 2026-01-01
    "%d-%m-%Y",   # 01-01-2026
    "%d/%m/%Y",   # 01/01/2026
    "%m/%d/%Y",   # 01/31/2026
    "%d-%m-%y",   # 01-01-26
    "%d/%m/%y",   # 01/01/26
]

def _parse_date(val) -> str:
    if not val:
        return date.today().isoformat()
    if isinstance(val, (date, datetime)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s.lower() in ("none", "nan", ""):
        return date.today().isoformat()
    for fmt in _DATE_FMTS:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return date.today().isoformat()


# ─── Column finder ────────────────────────────────────────────────────────────
def _cell_str(v) -> str:
    """Convert a raw Excel cell value to a clean string.
    Handles: None → '', float(9999999999.0) → '9999999999', datetime → ISO date.
    """
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float):
        # Remove spurious .0 so mobile/amount don't gain an extra digit
        if v == int(v):
            return str(int(v))
        return str(v)
    return str(v).strip()


def _col(row: list, header: list, *names) -> str:
    for n in names:
        if n in header:
            idx = header.index(n)
            if idx < len(row):
                v = _cell_str(row[idx]).strip()
                if v and v.lower() not in ("none", "nan"):
                    return v
    return ""


# ─── Core import logic ────────────────────────────────────────────────────────
def _import_rows(all_rows: list) -> dict:
    errors   = []
    imported = 0
    skipped  = 0

    if not all_rows:
        return {"imported": 0, "skipped": 0, "errors": ["Empty file"]}

    # Normalise header: lower-case, strip spaces and special chars
    raw_header  = all_rows[0]
    header      = [str(h).lower().strip().replace("'", "").replace("'", "") for h in raw_header]
    data_rows   = all_rows[1:]

    # Seed auto-receipt counter from last DB entry
    last         = db.get_next_receipt()
    auto_book    = last["book_no"]
    auto_receipt = last["receipt_no"]   # will be incremented before first use

    for i, row in enumerate(data_rows, 2):
        # Skip completely empty rows
        if not any(str(c).strip() for c in row):
            continue

        name   = _col(row, header,
                      "name", "member name", "fullname", "full name", "membername")
        mobile = _col(row, header,
                      "mobile", "phone", "mobile number", "contact", "mobile no",
                      "phone no", "phonenumber")

        if not name:
            errors.append(f"Row {i}: Missing name — skipped")
            skipped += 1
            continue
        if not mobile:
            errors.append(f"Row {i}: Missing mobile — skipped")
            skipped += 1
            continue

        # Clean mobile — digits only, must be 10 digits
        mobile_digits = "".join(c for c in mobile if c.isdigit())
        if len(mobile_digits) != 10:
            errors.append(f"Row {i}: {name} — mobile '{mobile}' is not 10 digits — skipped")
            skipped += 1
            continue

        # Dates
        join_raw   = _col(row, header,
                          "join date", "joindate", "joining date", "date of joining", "doj")
        expiry_raw = _col(row, header,
                          "expiry date", "expirydate", "expiry", "expiry dt",
                          "membership expiry", "end date", "valid till")

        join_date   = _parse_date(join_raw)   if join_raw   else date.today().isoformat()
        expiry_date = _parse_date(expiry_raw) if expiry_raw else None

        # Plan
        plan_raw    = _col(row, header, "plan", "membership", "membership plan",
                           "plan name", "duration")
        plan        = _normalise_plan(plan_raw) if plan_raw else "Monthly"

        # Amount
        amount_raw = _col(row, header, "amount", "fee", "fees", "paid", "payment")
        try:
            amount = float(str(amount_raw).replace(",", "")) if amount_raw else 0.0
        except ValueError:
            amount = 0.0

        # Father's Name (header may have smart-quote apostrophe)
        father = _col(row, header,
                      "fathers name", "father name", "father", "fathername",
                      "fathers name", "guardian")

        address = _col(row, header, "address", "addr")

        # Book No & Receipt No — use file values if present, else auto-increment
        book_raw    = _col(row, header, "book no", "bookno", "book", "book number")
        receipt_raw = _col(row, header, "receipt no", "receiptno", "receipt", "receipt number", "rcpt no")
        try:
            book_no = int(float(book_raw)) if book_raw else None
        except ValueError:
            book_no = None
        try:
            receipt_no = int(float(receipt_raw)) if receipt_raw else None
        except ValueError:
            receipt_no = None

        if book_no is None or receipt_no is None:
            # Auto-generate: continue from last receipt in DB
            book_no    = auto_book
            receipt_no = auto_receipt
        # Advance counter for next row (whether we used auto or file-supplied values)
        auto_book    = book_no
        auto_receipt = receipt_no + 1

        # If expiry_date not in file, calculate from plan + start_date
        if not expiry_date:
            from dateutil.relativedelta import relativedelta
            months     = _PLAN_MONTHS.get(plan, 1)
            start_dt   = datetime.strptime(join_date, "%Y-%m-%d").date()
            expiry_date = (start_dt + relativedelta(months=months)).isoformat()

        # Skip if same Name + Mobile already exists (duplicate guard)
        if db.member_exists(name, mobile_digits):
            errors.append(f"Row {i}: {name} ({mobile_digits}) already exists — skipped")
            skipped += 1
            continue

        try:
            db.add_member({
                "name": name, "mobile": mobile_digits,
                "father_name": father, "address": address,
                "join_date": join_date, "plan": plan,
                "amount": amount,
                "start_date": join_date,
                "expiry_date": expiry_date,
                "payment_mode": "Cash",
                "book_no": book_no,
                "receipt_no": receipt_no,
            })
            imported += 1
        except Exception as e:
            errors.append(f"Row {i}: {name} — {e}")
            skipped += 1

    return {"imported": imported, "skipped": skipped, "errors": errors[:100]}


# ─── Parse file → rows ────────────────────────────────────────────────────────
def _parse_file(content: bytes, filename: str) -> list:
    fname = filename.lower()
    if fname.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        # Keep raw values (datetime, int, float, str) — _col/_cell_str handles type conversion
        return [list(row) for row in ws.iter_rows(values_only=True)]

    elif fname.endswith(".ods"):
        from odf.opendocument import load as ods_load
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        doc   = ods_load(io.BytesIO(content))
        sheet = doc.spreadsheet.getElementsByType(Table)[0]
        rows  = []
        for tr in sheet.getElementsByType(TableRow):
            row = []
            for tc in tr.getElementsByType(TableCell):
                ps = tc.getElementsByType(P)
                row.append(str(ps[0]) if ps else "")
            rows.append(row)
        return rows

    elif fname.endswith(".csv"):
        text = content.decode("utf-8-sig")
        return list(csv.reader(io.StringIO(text)))

    return []


# ─── Routes ───────────────────────────────────────────────────────────────────

@router.get("")
async def import_export_page(request: Request):
    return templates.TemplateResponse("import_export.html", {"request": request})


@router.get("/template/xlsx")
async def download_template_xlsx():
    """Generate and download a sample XLSX template matching the expected format."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"

    headers = ["Name", "Join Date", "Expiry Date", "Plan", "Mobile", "Father's Name", "Amount", "Book No", "Receipt No"]
    col_widths = [20, 14, 14, 14, 14, 22, 10, 10, 12]
    hdr_fill = PatternFill("solid", fgColor="1a6b3a")
    hdr_font = Font(bold=True, color="FFFFFF")

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    # 3 sample rows matching the user's actual format
    samples = [
        ("Rahul Sharma",  "1-Jan-26", "31-Mar-26", "3 Months",  "9876543210", "Suresh Sharma", 2100, 1, 1),
        ("Priya Singh",   "1-Jan-26", "31-Jan-26", "1 Months",  "9876543211", "Mohan Singh",    800, 1, 2),
        ("Amit Kumar",    "1-Jan-26", "30-Jun-26", "6 Months",  "9876543212", "Vijay Kumar",   5000, 1, 3),
    ]
    for sample in samples:
        ws.append(sample)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=member_import_template.xlsx"},
    )


@router.get("/template/csv")
async def download_template_csv():
    """Generate and download a sample CSV template."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Name", "Join Date", "Expiry Date", "Plan", "Mobile", "Father's Name", "Amount", "Book No", "Receipt No"])
    writer.writerows([
        ["Rahul Sharma", "1-Jan-26", "31-Mar-26", "3 Months", "9876543210", "Suresh Sharma", 2100, 1, 1],
        ["Priya Singh",  "1-Jan-26", "31-Jan-26", "1 Months", "9876543211", "Mohan Singh",    800, 1, 2],
        ["Amit Kumar",   "1-Jan-26", "30-Jun-26", "6 Months", "9876543212", "Vijay Kumar",   5000, 1, 3],
    ])
    buf = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=member_import_template.csv"},
    )


@router.post("/preview")
async def preview_file(file: UploadFile = File(...)):
    """Parse uploaded file and return first 5 data rows for preview."""
    content = await file.read()
    try:
        rows = _parse_file(content, file.filename)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=400)

    if not rows:
        return JSONResponse({"error": "File is empty or unreadable."}, status_code=400)

    # Filter out completely empty rows
    rows = [r for r in rows if any(str(c).strip() for c in r)]

    header    = [_cell_str(c) for c in (rows[0] if rows else [])]
    data_rows = [[_cell_str(c) for c in row] for row in rows[1:6]]

    return JSONResponse({
        "header":     header,
        "rows":       data_rows,
        "total_rows": len(rows) - 1,
        "filename":   file.filename,
    })


@router.post("/import")
async def import_file(file: UploadFile = File(...)):
    content  = await file.read()
    filename = file.filename.lower()

    if not any(filename.endswith(ext) for ext in (".xlsx", ".ods", ".csv")):
        return JSONResponse({"error": "Unsupported file. Use XLSX, ODS, or CSV."}, status_code=400)

    try:
        rows   = _parse_file(content, filename)
        result = _import_rows(rows)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

    return JSONResponse(result)
