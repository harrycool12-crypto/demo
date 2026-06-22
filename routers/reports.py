import io
import os
from datetime import date
from fastapi import APIRouter, Request, Query
from fastapi.responses import StreamingResponse, JSONResponse
import database as db
from config import templates

router = APIRouter(prefix="/reports", tags=["reports"])


@router.get("/active")
async def active_members(request: Request, status: str = "Active"):
    if status == "All":
        members = db.report_all_members()
    elif status == "Inactive":
        members = db.report_inactive_members_full()
    else:
        members = db.report_active_members()
    return templates.TemplateResponse("reports/active.html", {
        "request": request, "members": members, "status": status
    })


@router.get("/inactive")
async def inactive_members(request: Request):
    members = db.report_inactive_members()
    return templates.TemplateResponse("reports/inactive.html", {"request": request, "members": members})


@router.get("/expiry")
async def expiry_report(request: Request, days: int = 30):
    members = db.report_expiry(days)
    return templates.TemplateResponse("reports/expiry.html", {
        "request": request, "members": members, "days": days
    })


@router.get("/collection")
async def collection_report(
    request: Request,
    period: str = "monthly",
    year: int = None,
    month: int = None,
):
    today = date.today()
    year  = year  or today.year
    month = month or today.month
    data  = db.report_collection(period, year, month)
    return templates.TemplateResponse("reports/collection.html", {
        "request": request, "data": data, "period": period, "year": year, "month": month,
        "years": list(range(today.year - 3, today.year + 1)),
        "months": [(i, date(2000, i, 1).strftime("%B")) for i in range(1, 13)],
    })


# ─── Excel export helpers ─────────────────────────────────────────────────────

def _excel_response(wb, filename: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/active")
async def export_active(status: str = "Active"):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    if status == "All":
        members = db.report_all_members()
    elif status == "Inactive":
        members = db.report_inactive_members_full()
    else:
        members = db.report_active_members()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{status} Members"
    show_status = (status in ("All", "Inactive"))
    headers = ["Receipt No", "Member ID", "Name", "Mobile", "Father's Name", "Address", "Join Date", "Plan", "Expiry Date"]
    col_widths = [12, 12, 25, 15, 25, 35, 14, 14, 14]
    if show_status:
        headers.append("Status")
        col_widths.append(12)
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a6b3a")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 20
    for m in members:
        rcpt = f"B{m['book_no']}-{m['receipt_no']:03d}" if m.get("receipt_no") else ""
        row = [
            rcpt, m["member_id"], m["name"], m["mobile"],
            m.get("father_name", "") or "", m.get("address", "") or "",
            m["join_date"], m.get("plan", "") or "", m.get("expiry_date", "") or "",
        ]
        if show_status:
            row.append(m.get("status", ""))
        ws.append(row)
    return _excel_response(wb, f"{status.lower()}_members.xlsx")


@router.get("/export/inactive")
async def export_inactive():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    members = db.report_inactive_members()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inactive Members"
    headers = ["Member ID", "Name", "Mobile", "Father's Name", "Address", "Join Date", "Last Expiry"]
    col_widths = [12, 25, 15, 25, 35, 14, 14]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a6b3a")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 20
    for m in members:
        ws.append([
            m["member_id"], m["name"], m["mobile"],
            m.get("father_name", "") or "", m.get("address", "") or "",
            m["join_date"], m.get("last_expiry", "") or "",
        ])
    return _excel_response(wb, "inactive_members.xlsx")


@router.get("/export/expiry")
async def export_expiry(days: int = 30):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    members = db.report_expiry(days)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expiry Report"
    headers = ["Member ID", "Name", "Mobile", "Plan", "Expiry Date", "Days Remaining"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a6b3a")
    for m in members:
        ws.append([m["member_id"], m["name"], m["mobile"], m.get("plan", ""), m["expiry_date"], m["days_remaining"]])
    return _excel_response(wb, "expiry_report.xlsx")


@router.get("/export/payment-history")
async def export_payment_history(
    member_id: int = None,
    month: str = "",
    start_date: str = "",
    end_date: str = "",
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    payments = db.get_payment_history(
        member_id, month or None, start_date or None, end_date or None
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment History"
    headers = ["Receipt No", "Member Name", "Plan", "Start Date", "Expiry Date",
               "Amount (₹)", "Payment Date", "Mode"]
    col_widths = [12, 25, 14, 14, 14, 14, 14, 14]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a6b3a")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 20
    total = 0
    for p in payments:
        rcpt = f"B{p['book_no']}-{p['receipt_no']:03d}" if p.get("receipt_no") else ""
        ws.append([
            rcpt, p["name"], p.get("plan", "") or "",
            p.get("membership_start", "") or "",
            p.get("membership_expiry", "") or "",
            p["amount"], p["payment_date"],
            p.get("payment_mode", "") or "",
        ])
        total += p["amount"]
    # Total row
    from openpyxl.styles import Font as F
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=5, value="TOTAL").font = F(bold=True)
    ws.cell(row=total_row, column=6, value=total).font = F(bold=True)
    return _excel_response(wb, "payment_history.xlsx")


@router.get("/export/collection")
async def export_collection(period: str = "monthly", year: int = None, month: int = None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    today = date.today()
    year  = year  or today.year
    month = month or today.month
    data  = db.report_collection(period, year, month)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Collection Report"
    headers = ["Period", "Amount (INR)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a6b3a")
    for row in data["rows"]:
        ws.append([row["label"], row["amount"]])
    ws.append(["TOTAL", data["total"]])
    return _excel_response(wb, "collection_report.xlsx")


@router.get("/export/all-members")
async def export_all_members():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    members = db.search_members()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Members"
    headers = ["Member ID", "Name", "Mobile", "Father's Name", "Address", "Join Date", "Status", "Plan", "Expiry Date"]
    col_widths = [12, 25, 15, 25, 35, 14, 12, 14, 14]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a6b3a")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 20
    for m in members:
        ws.append([
            m["member_id"], m["name"], m["mobile"],
            m.get("father_name", "") or "", m.get("address", "") or "",
            m["join_date"], m["status"], m.get("plan", "") or "", m.get("expiry_date", "") or "",
        ])
    return _excel_response(wb, "all_members.xlsx")
